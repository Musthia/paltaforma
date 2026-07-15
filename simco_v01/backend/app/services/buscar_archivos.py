import os
from typing import Optional

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")


def _texto_plano(ruta: str) -> Optional[str]:
    try:
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception:
        return None


def _docx(ruta: str) -> Optional[str]:
    try:
        from docx import Document
        doc = Document(ruta)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        return None


def _xlsx(ruta: str) -> Optional[str]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta, read_only=True, data_only=True)
        textos = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                textos.append(" ".join(str(c) for c in row if c is not None))
        wb.close()
        return "\n".join(textos)
    except Exception:
        return None


def _pdf(ruta: str) -> Optional[str]:
    try:
        import fitz
        doc = fitz.open(ruta)
        texto = "\n".join(page.get_text() for page in doc)
        doc.close()
        return texto
    except Exception:
        return None


def _rtf(ruta: str) -> Optional[str]:
    try:
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            raw = f.read()
        # Strip simple RTF: remove everything between { } and tags
        import re
        text = re.sub(r"\\.*?[ {]", " ", raw)
        text = re.sub(r"[{}\\]", " ", text)
        return text.strip()
    except Exception:
        return None


READERS = {
    ".txt": _texto_plano,
    ".csv": _texto_plano,
    ".log": _texto_plano,
    ".md": _texto_plano,
    ".json": _texto_plano,
    ".xml": _texto_plano,
    ".ini": _texto_plano,
    ".cfg": _texto_plano,
    ".yaml": _texto_plano,
    ".yml": _texto_plano,
    ".html": _texto_plano,
    ".htm": _texto_plano,
    ".css": _texto_plano,
    ".js": _texto_plano,
    ".py": _texto_plano,
    ".sql": _texto_plano,
    ".rtf": _rtf,
    ".docx": _docx,
    ".doc": _docx,
    ".xlsx": _xlsx,
    ".xls": _xlsx,
    ".pdf": _pdf,
}


def leer_archivo(nombre_servidor: str) -> Optional[str]:
    ruta = os.path.join(UPLOAD_DIR, nombre_servidor)
    if not os.path.exists(ruta):
        return None
    ext = os.path.splitext(nombre_servidor)[1].lower()
    reader = READERS.get(ext)
    if not reader:
        return None
    return reader(ruta)


def buscar_en_contenido(nombre_archivo_db: str, query: str) -> Optional[str]:
    if not nombre_archivo_db or "::" not in nombre_archivo_db:
        return None
    nombre_servidor, nombre_original = nombre_archivo_db.split("::", 1)
    texto = leer_archivo(nombre_servidor)
    if texto is None:
        return None
    q = query.lower()
    if q in texto.lower():
        return nombre_original
    return None
