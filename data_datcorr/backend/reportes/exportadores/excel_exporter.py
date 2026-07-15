import logging
import io
from datetime import datetime

from fastapi.responses import StreamingResponse

logger = logging.getLogger("datcorr")


class ExcelExporter:

    FORMATO = "xlsx"

    def exportar(self, datos: list[dict], nombre_archivo: str) -> StreamingResponse:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = nombre_archivo[:31]

            if datos:
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")

                for col_idx, key in enumerate(datos[0].keys(), 1):
                    cell = ws.cell(row=1, column=col_idx, value=key)
                    cell.font = header_font
                    cell.fill = header_fill

                for row_idx, row in enumerate(datos, 2):
                    for col_idx, key in enumerate(row.keys(), 1):
                        val = row[key]
                        if isinstance(val, datetime):
                            val = val.replace(tzinfo=None)
                        ws.cell(row=row_idx, column=col_idx, value=val)

                ws.auto_filter.ref = ws.dimensions

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{nombre_archivo}.xlsx"',
                    "Cache-Control": "private",
                },
            )
        except Exception as e:
            logger.error(f"Error generando Excel: {e}", exc_info=True)
            raise
